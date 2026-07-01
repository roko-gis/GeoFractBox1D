"""
GeoFracBox1D - Fractal Box-Counting Analysis for QGIS
Version 1.0

Empirical box-counting estimator of fractal scaling behavior in 1D datasets
with statistical diagnostics for power-law validity assessment.

References:
    Mandelbrot (1982), Turcotte (1986), Ranguelov et al. (2002-2019)
"""

from qgis.PyQt.QtWidgets import (QDialog, QVBoxLayout, QHBoxLayout, QPushButton,
                                  QLabel, QComboBox, QTableWidget, QTableWidgetItem,
                                  QFileDialog, QMessageBox, QGroupBox, QCheckBox,
                                  QProgressBar, QApplication, QTextEdit, QDialogButtonBox,
                                  QFrame, QSplitter, QWidget, QMenu, QAction)
from qgis.PyQt.QtCore import Qt
from qgis.PyQt.QtGui import QColor, QCursor
from qgis.core import QgsVectorLayer, QgsProject, QgsFeatureRequest

import json
import os
from datetime import datetime

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
from matplotlib.patches import Rectangle
from scipy import stats


# ============================================================================
# PLUGIN METADATA
# ============================================================================
PLUGIN_VERSION = "1.0"
PLUGIN_METHOD = "GeoFracBox1D"
SURROGATE_SEED = 42

# ============================================================================
# NUMERICAL THRESHOLDS
# ============================================================================
EPS = 1e-12

# R² thresholds
R2_EXCELLENT = 0.95
R2_GOOD = 0.90
R2_ACCEPTABLE = 0.85
R2_REGION_MIN = 0.70

# Sample requirements
MIN_TOTAL_POINTS = 10
MIN_SAMPLE_SIZE = 50
MIN_SCALING_POINTS = 4
MIN_VALID_SCALES = 6
MIN_SURROGATES = 5

# Scale span thresholds (decades)
SCALE_SPAN_EXCELLENT = 2.0
SCALE_SPAN_WEAK = 1.5
SCALE_SPAN_MIN = 1.0

# Curvature thresholds
CURVATURE_HIGH = 0.5
CURVATURE_MODERATE = 0.3

# Slope stability thresholds
SLOPE_UNSTABLE = 0.3
SLOPE_WARNING = 0.15

# Confidence interval
CI_WIDE = 0.5

# Surrogate test
SURROGATE_SIGNIFICANCE = 0.05
SURROGATE_MARGINAL = 0.01

# Scaling fraction
SCALING_FRACTION_MIN = 0.5
SCALING_FRACTION_CRITICAL = 0.3

# Confidence scores
CONFIDENCE_MARGINAL_LOW = 0.5
CONFIDENCE_MARGINAL_HIGH = 0.6
CONFIDENCE_WEAK = 0.7
CONFIDENCE_STRONG = 0.7

# Scale counts
FAST_SCALES = 12
STANDARD_SCALES = 20
HIGH_PRECISION_SCALES = 35
SURROGATE_SCALES = 15

# Plot settings
PLOT_DPI = 300
MAX_LABELS_BEFORE_SKIP = 15
LABEL_FONT_REGION = 8
LABEL_FONT_OTHER = 7

# ============================================================================
# SEVERITY SYSTEM
# ============================================================================
SEVERITY_CRITICAL = 3
SEVERITY_WARNING = 2
SEVERITY_INFO = 1
SEVERITY_OK = 0

SEVERITY_CONFIG = {
    SEVERITY_CRITICAL: ("🔴", "#cc0000"),
    SEVERITY_WARNING: ("🟡", "#cc8800"),
    SEVERITY_INFO: ("🔵", "#3366cc"),
    SEVERITY_OK: ("⚪", "#666666"),
}

# ============================================================================
# STATUS CONFIGURATION
# ============================================================================
STATUS_CONFIG = {
    "STRONG POWER-LAW": {
        "icon": "✓",
        "box_color": "lightgreen",
        "bg": Qt.green,
        "style": "font-weight: bold; color: green; padding: 4px; font-size: 11px;",
    },
    "WEAK POWER-LAW": {
        "icon": "⚠",
        "box_color": "lightyellow",
        "bg": Qt.yellow,
        "style": (
            "font-weight: bold; color: #cc8800; background-color: #fff8e6; "
            "padding: 4px; font-size: 11px;"
        ),
    },
    "MARGINAL POWER-LAW": {
        "icon": "~",
        "box_color": "navajowhite",
        "bg": QColor("#ffb347"),
        "style": (
            "font-weight: bold; color: #b8600c; background-color: #fff3cd; "
            "padding: 4px; font-size: 11px;"
        ),
    },
    "NOT POWER-LAW": {
        "icon": "✗",
        "box_color": "lightcoral",
        "bg": Qt.red,
        "style": (
            "font-weight: bold; color: white; background-color: #cc0000; "
            "padding: 4px; font-size: 11px;"
        ),
    },
}

# ============================================================================
# UI STYLESHEETS
# ============================================================================
STYLE_VERSION_LABEL = """
    QLabel {
        background-color: #2c3e50; color: white; padding: 4px 10px;
        border-radius: 3px; font-size: 11px; font-weight: bold;
    }
"""

STYLE_METHODOLOGY_BTN = """
    QPushButton {
        background-color: #e8f0fe; border: 1px solid #1967d2;
        border-radius: 4px; padding: 6px 12px; font-size: 11px;
        font-weight: bold; color: #1967d2;
    }
    QPushButton:hover { background-color: #d2e3fc; }
"""

STYLE_RUN_BTN = """
    QPushButton {
        font-weight: bold; font-size: 13px; background-color: #1967d2;
        color: white; border: none; border-radius: 4px; padding: 8px;
    }
    QPushButton:hover { background-color: #1557b0; }
    QPushButton:disabled { background-color: #ccc; color: #888; }
"""

STYLE_BADGE_GOOD = (
    "QLabel { background-color: #d4edda; color: #155724; "
    "border: 1px solid #c3e6cb; border-radius: 3px; "
    "padding: 2px 5px; font-size: 9px; font-weight: bold; }"
)
STYLE_BADGE_WARNING = (
    "QLabel { background-color: #fff3cd; color: #856404; "
    "border: 1px solid #ffeeba; border-radius: 3px; "
    "padding: 2px 5px; font-size: 9px; font-weight: bold; }"
)
STYLE_BADGE_BAD = (
    "QLabel { background-color: #f8d7da; color: #721c24; "
    "border: 1px solid #f5c6cb; border-radius: 3px; "
    "padding: 2px 5px; font-size: 9px; font-weight: bold; }"
)
STYLE_BADGE_NEUTRAL = (
    "QLabel { background-color: #e2e3e5; color: #383d41; "
    "border: 1px solid #d6d8db; border-radius: 3px; "
    "padding: 2px 5px; font-size: 9px; }"
)


# ============================================================================
# METHODOLOGY DIALOG
# ============================================================================
class MethodologyDialog(QDialog):
    """Dialog displaying methodology, assumptions, limitations and references."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("GeoFracBox1D - Methodology & Theoretical Background")
        self.setMinimumSize(750, 650)
        self.setModal(False)
        self._setup_ui()

    def _setup_ui(self):
        layout = QVBoxLayout()

        self.text_edit = QTextEdit()
        self.text_edit.setReadOnly(True)
        self.text_edit.setStyleSheet(
            "QTextEdit { background-color: #ffffff; border: 1px solid #dee2e6; "
            "border-radius: 4px; padding: 12px; font-size: 12px; line-height: 1.5; }"
        )
        self.text_edit.setHtml(self._build_html())
        layout.addWidget(self.text_edit)

        button_box = QDialogButtonBox(QDialogButtonBox.Ok)
        button_box.accepted.connect(self.accept)
        layout.addWidget(button_box)
        self.setLayout(layout)

    @staticmethod
    def _build_html():
        return """
        <div style="font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, Arial, sans-serif;">
        <h2 style="margin-top:0;color:#2c3e50;">GeoFracBox1D — Fractal Box-Counting Analysis</h2>
        <p><b>GeoFracBox1D</b> provides an <b>empirical box-counting estimator</b> of fractal
        scaling behavior in 1D datasets, with statistical diagnostics for power-law validity.
        <b>Note:</b> D is a box-counting empirical scaling exponent, not an invariant constant.</p>
        <h3 style="color:#2c3e50;">Box-Counting Principle</h3>
        <p style="text-align:center;font-size:16px;font-weight:bold;margin:15px 0;
        background-color:#f0f4f8;padding:10px;border-radius:4px;">N(ε) ∝ ε<sup>−D</sup></p>
        <ul><li><b>N(ε)</b> — occupied intervals at scale ε</li>
        <li><b>ε</b> — scale size</li><li><b>D</b> — empirical scaling exponent (0 ≤ D ≤ 1)</li></ul>
        <h3 style="color:#2c3e50;">Diagnostics</h3>
        <ul><li><b>R²</b> — goodness of fit</li><li><b>D ± CI</b> — 95% confidence interval</li>
        <li><b>Curvature test</b> — detects non-power-law behavior</li>
        <li><b>Surrogate test</b> — permutation-based significance</li>
        <li><b>Slope stability</b> — sliding-window consistency</li></ul>
        <h3 style="color:#2c3e50;">Assumptions and Limitations</h3>
        <div style="background-color:#fff9e6;border-left:4px solid #f0c040;padding:10px;
        margin:10px 0;border-radius:4px;"><ul style="margin:0;padding-left:18px;">
        <li>Assumes <b>scale-invariant behavior</b> over at least part of the range.</li>
        <li>Results are <b>sensitive to noise and sampling density</b>.</li>
        <li>D is an <b>empirical approximation</b>, not a universal constant.</li>
        <li>Interpretation requires consideration of <b>all diagnostic indicators</b>.</li>
        </ul></div>
        <h3 style="color:#2c3e50;">References</h3>
        <div style="font-size:11px;line-height:1.4;color:#555;">
        <p>[1] Mandelbrot, B.B. (1982). <i>The Fractal Geometry of Nature.</i> Freeman, New York.</p>
        <p>[2] Turcotte, D.L. (1986). Fractals and fragmentation. <i>JGR</i>, 91(B2), 1921–1926.</p>
        <p>[3] Ranguelov, B., & Dimitrova, S. (2002). Fractal model of the recent surface
        earth crust fragmentation in Bulgaria. <i>Comptes Rendus BAS</i>, 55(3).</p>
        <p>[4] Ranguelov, B., Dimitrova, S., Gospodinov, D., et al. (2004). Fractal properties
        of the South Balkans seismotectonic model. <i>Proc. 5th ISEMG</i> (pp. 643–646).</p>
        <p>[5] Ranguelov, B. (2010). Nonlinearities and fractal properties of the
        European–Mediterranean seismotectonic model. <i>Geodynamics & Tectonophysics</i>, 1(3).</p>
        <p>[6] Ranguelov, B., & Ivanov, Y. (2017). Fractal properties of plate tectonics.
        <i>J. Mining Geol. Sci.</i>, 60(1), 83–89.</p>
        <p>[7] Ranguelov, B., & Shadiya, F. (2018). Fractals, natural disasters and ecological
        problems of Maldives. <i>Ecol. Eng. Environ. Prot.</i>, 2, 18–25.</p>
        <p>[8] Ranguelov, B., Shadiya, F., & Ivanov, Y. (2018). Fractal nature of the
        Maldives Arc. <i>Proc. SGEM 2018</i> (pp. 81–86).</p>
        <p>[9] Ranguelov, B., Iliev, R., Tzankov, T., & Spassov, E. (2019). Fractal analysis
        of the lunar free-air gravity field. <i>To Physics Journal</i>, 2, 126–133.</p>
        </div></div>
        """


# ============================================================================
# MAIN ANALYSIS DIALOG
# ============================================================================
class FractalAnalysisDialog(QDialog):
    """Main dialog for GeoFracBox1D fractal box-counting analysis."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("GeoFracBox1D - Fractal Analysis")
        self.setMinimumSize(1400, 820)
        self.setModal(False)

        self.layer = None
        self.results = None
        self.external_data = None
        self._methodology_dialog = None
        self.timestamp = None

        self._setup_ui()
        self.refresh_layers()
        self.show()

    # ========================================================================
    # MESSAGING HELPERS
    # ========================================================================
    def _show_error(self, text):
        QMessageBox.critical(self, "Error", text)

    def _show_warning(self, title, text):
        QMessageBox.warning(self, title, text)

    def _show_info(self, title, text):
        QMessageBox.information(self, title, text)

    # ========================================================================
    # UI BUILDERS
    # ========================================================================
    def _setup_ui(self):
        splitter = QSplitter(Qt.Horizontal)
        splitter.addWidget(self._build_left_panel())
        splitter.addWidget(self._build_right_panel())
        splitter.setSizes([440, 780])

        layout = QVBoxLayout()
        layout.setContentsMargins(4, 4, 4, 4)
        layout.addWidget(splitter)
        self.setLayout(layout)

    def _build_left_panel(self):
        panel = QWidget()
        panel.setMinimumWidth(420)
        panel.setMaximumWidth(520)
        layout = QVBoxLayout()
        layout.setContentsMargins(8, 8, 8, 8)
        layout.setSpacing(5)

        layout.addLayout(self._build_header())
        layout.addWidget(self._build_mode_group())
        layout.addWidget(self._build_source_group())
        layout.addWidget(self._build_attribute_group())
        layout.addWidget(self._build_quality_frame())
        layout.addWidget(self._progress_bar)
        layout.addWidget(self._status_label)
        layout.addWidget(self._build_results_group())
        layout.addWidget(self._build_plot_controls())
        layout.addWidget(self._build_export_frame())
        layout.addStretch()

        panel.setLayout(layout)
        return panel

    def _build_header(self):
        bar = QHBoxLayout()

        self.version_label = QLabel(f" v{PLUGIN_VERSION}")
        self.version_label.setStyleSheet(STYLE_VERSION_LABEL)

        self.methodology_btn = QPushButton("📚 Methodology")
        self.methodology_btn.setStyleSheet(STYLE_METHODOLOGY_BTN)
        self.methodology_btn.clicked.connect(self._show_methodology)

        bar.addWidget(self.version_label)
        bar.addWidget(self.methodology_btn)
        bar.addStretch()
        return bar

    def _build_mode_group(self):
        group = QGroupBox("Analysis Mode")
        layout = QHBoxLayout()
        layout.addWidget(QLabel("Scales:"))
        self.scale_mode_combo = QComboBox()
        self.scale_mode_combo.addItems(
            ["Fast (12)", "Standard (20)", "High precision (35)"]
        )
        self.scale_mode_combo.setCurrentIndex(1)
        layout.addWidget(self.scale_mode_combo)
        layout.addStretch()
        group.setLayout(layout)
        return group

    def _build_source_group(self):
        group = QGroupBox("Data Source")
        layout = QVBoxLayout()
        layout.setSpacing(3)

        row1 = QHBoxLayout()
        row1.addWidget(QLabel("Layer:"))
        self.layer_combo = QComboBox()
        self.layer_combo.setMinimumWidth(140)
        self.layer_combo.currentIndexChanged.connect(self._on_layer_changed)
        self.refresh_btn = QPushButton("↻")
        self.refresh_btn.setMaximumWidth(35)
        self.refresh_btn.clicked.connect(self.refresh_layers)
        row1.addWidget(self.layer_combo)
        row1.addWidget(self.refresh_btn)

        row2 = QHBoxLayout()
        self.load_file_btn = QPushButton("📂 Open File")
        self.load_file_btn.clicked.connect(self._load_file)
        self.file_label = QLabel("No file loaded")
        self.file_label.setStyleSheet("color: gray; font-size: 10px;")
        row2.addWidget(self.load_file_btn)
        row2.addWidget(self.file_label)
        row2.addStretch()

        layout.addLayout(row1)
        layout.addLayout(row2)
        group.setLayout(layout)
        return group

    def _build_attribute_group(self):
        group = QGroupBox("Attribute")
        layout = QVBoxLayout()
        layout.setSpacing(3)

        row = QHBoxLayout()
        row.addWidget(QLabel("Field:"))
        self.attr_combo = QComboBox()
        self.attr_combo.setMinimumWidth(140)
        self.attr_combo.currentIndexChanged.connect(self._on_attr_changed)
        row.addWidget(self.attr_combo)
        row.addStretch()

        self.info_label = QLabel("Samples: -- | Range: --")
        self.info_label.setStyleSheet("font-size: 10px; color: #666;")

        self.run_btn = QPushButton("▶ Run Analysis")
        self.run_btn.clicked.connect(self.run_analysis)
        self.run_btn.setEnabled(False)
        self.run_btn.setMinimumHeight(36)
        self.run_btn.setStyleSheet(STYLE_RUN_BTN)

        layout.addLayout(row)
        layout.addWidget(self.info_label)
        layout.addWidget(self.run_btn)
        group.setLayout(layout)
        return group

    def _build_quality_frame(self):
        frame = QFrame()
        frame.setFrameStyle(QFrame.StyledPanel)
        frame.setStyleSheet(
            "QFrame { background-color: #f8f9fa; border-radius: 4px; }"
        )
        layout = QHBoxLayout()
        layout.setContentsMargins(4, 3, 4, 3)
        layout.setSpacing(3)
        layout.addWidget(QLabel("Quality:"))

        self.badge_sample = self._make_badge("Samples", 70)
        self.badge_span = self._make_badge("Span", 60)
        self.badge_scales = self._make_badge("Scales", 60)

        layout.addWidget(self.badge_sample)
        layout.addWidget(self.badge_span)
        layout.addWidget(self.badge_scales)
        layout.addStretch()
        frame.setLayout(layout)
        return frame

    @staticmethod
    def _make_badge(text, width):
        badge = QLabel(text)
        badge.setAlignment(Qt.AlignCenter)
        badge.setMinimumWidth(width)
        badge.setStyleSheet(STYLE_BADGE_NEUTRAL)
        return badge

    def _build_results_group(self):
        group = QGroupBox("Results")
        layout = QVBoxLayout()
        layout.setContentsMargins(3, 6, 3, 3)

        self.results_table = QTableWidget()
        self.results_table.setColumnCount(2)
        self.results_table.setHorizontalHeaderLabels(["Parameter", "Value"])
        self.results_table.horizontalHeader().setStretchLastSection(True)
        self.results_table.setAlternatingRowColors(True)
        self.results_table.setMaximumHeight(250)
        layout.addWidget(self.results_table)
        group.setLayout(layout)
        return group

    def _build_plot_controls(self):
        frame = QFrame()
        frame.setFrameStyle(QFrame.StyledPanel)
        frame.setStyleSheet(
            "QFrame { background-color: #fafafa; border-radius: 4px; }"
        )
        layout = QVBoxLayout()
        layout.setContentsMargins(6, 3, 6, 3)
        layout.setSpacing(2)

        row1 = QHBoxLayout()
        self.show_labels_check = QCheckBox("N labels")
        self.show_labels_check.setChecked(True)
        self.show_grid_check = QCheckBox("Grid")
        self.show_grid_check.setChecked(True)
        row1.addWidget(self.show_labels_check)
        row1.addWidget(self.show_grid_check)
        row1.addStretch()

        row2 = QHBoxLayout()
        self.show_regression_check = QCheckBox("Regression")
        self.show_regression_check.setChecked(True)
        self.show_all_points_check = QCheckBox("All scales")
        self.show_all_points_check.setChecked(True)
        row2.addWidget(self.show_regression_check)
        row2.addWidget(self.show_all_points_check)
        row2.addStretch()

        self.update_plot_btn = QPushButton("↻ Refresh Plot")
        self.update_plot_btn.clicked.connect(self.update_plot)
        self.update_plot_btn.setEnabled(False)
        self.update_plot_btn.setMaximumHeight(26)

        layout.addLayout(row1)
        layout.addLayout(row2)
        layout.addWidget(self.update_plot_btn)
        frame.setLayout(layout)
        return frame

    def _build_export_frame(self):
        frame = QFrame()
        frame.setFrameStyle(QFrame.StyledPanel)
        frame.setStyleSheet(
            "QFrame { background-color: #fafafa; border-radius: 4px; }"
        )
        layout = QHBoxLayout()
        layout.setContentsMargins(6, 4, 6, 4)
        layout.setSpacing(4)

        layout.addWidget(QLabel("Export:"))

        self.export_format_combo = QComboBox()
        self.export_format_combo.addItems(
            ["Excel (.xlsx)", "CSV (.csv)", "JSON (.json)"]
        )

        self.export_btn = QPushButton("📥 Export")
        self.export_btn.clicked.connect(self.export_result)
        self.export_btn.setEnabled(False)
        self.export_btn.setMaximumHeight(28)

        self.export_png_btn = QPushButton("📊 PNG")
        self.export_png_btn.clicked.connect(lambda: self._export_plot("png"))
        self.export_png_btn.setEnabled(False)
        self.export_png_btn.setMaximumHeight(28)

        self.export_jpg_btn = QPushButton("🖼 JPG")
        self.export_jpg_btn.clicked.connect(lambda: self._export_plot("jpg"))
        self.export_jpg_btn.setEnabled(False)
        self.export_jpg_btn.setMaximumHeight(28)

        self.export_pdf_btn = QPushButton("📄 PDF")
        self.export_pdf_btn.clicked.connect(lambda: self._export_plot("pdf"))
        self.export_pdf_btn.setEnabled(False)
        self.export_pdf_btn.setMaximumHeight(28)

        self.export_clipboard_btn = QPushButton("📋")
        self.export_clipboard_btn.clicked.connect(self.copy_to_clipboard)
        self.export_clipboard_btn.setEnabled(False)
        self.export_clipboard_btn.setMaximumHeight(28)
        self.export_clipboard_btn.setMaximumWidth(35)

        layout.addWidget(self.export_format_combo)
        layout.addWidget(self.export_btn)
        layout.addWidget(self.export_png_btn)
        layout.addWidget(self.export_jpg_btn)
        layout.addWidget(self.export_pdf_btn)
        layout.addWidget(self.export_clipboard_btn)
        layout.addStretch()
        frame.setLayout(layout)
        return frame

    def _build_right_panel(self):
        panel = QWidget()
        layout = QVBoxLayout()
        layout.setContentsMargins(4, 8, 8, 8)

        group = QGroupBox("Fractal Analysis Plot")
        group_layout = QVBoxLayout()
        group_layout.setContentsMargins(4, 4, 4, 4)

        self.figure = Figure(figsize=(9, 7))
        self.canvas = FigureCanvas(self.figure)
        self.canvas.setMinimumSize(550, 500)
        group_layout.addWidget(self.canvas)
        group.setLayout(group_layout)

        layout.addWidget(group)
        panel.setLayout(layout)
        return panel

    # ========================================================================
    # LAZY UI COMPONENTS
    # ========================================================================
    @property
    def _progress_bar(self):
        if not hasattr(self, "__progress_bar"):
            self.__progress_bar = QProgressBar()
            self.__progress_bar.setVisible(False)
            self.__progress_bar.setMaximumHeight(16)
        return self.__progress_bar

    @property
    def _status_label(self):
        if not hasattr(self, "__status_label"):
            self.__status_label = QLabel("Ready")
            self.__status_label.setStyleSheet(
                "font-weight: bold; color: blue; padding: 3px; font-size: 11px;"
            )
            self.__status_label.setWordWrap(True)
        return self.__status_label

    # ========================================================================
    # BADGE HELPERS
    # ========================================================================
    @staticmethod
    def _badge_style(status):
        styles = {
            "good": STYLE_BADGE_GOOD,
            "warning": STYLE_BADGE_WARNING,
            "bad": STYLE_BADGE_BAD,
        }
        return styles.get(status, STYLE_BADGE_NEUTRAL)

    def _update_badges(self, n_samples, scale_span, n_scales):
        self._set_badge(
            self.badge_sample, n_samples, 100, MIN_SAMPLE_SIZE, "Samples"
        )
        self._set_badge(
            self.badge_span,
            scale_span,
            SCALE_SPAN_EXCELLENT,
            SCALE_SPAN_MIN,
            "Span",
            is_span=True,
        )
        self._set_badge(
            self.badge_scales, n_scales, 10, MIN_VALID_SCALES, "Scales"
        )

    def _set_badge(self, badge, value, excellent, minimum, label, is_span=False):
        if value >= excellent:
            badge.setText(f"{label} ✓")
            badge.setStyleSheet(self._badge_style("good"))
        elif value >= minimum:
            text = f"~{value:.1f}" if is_span else f"~{value}"
            badge.setText(text)
            badge.setStyleSheet(self._badge_style("warning"))
        else:
            badge.setText(f"{label} ✗")
            badge.setStyleSheet(self._badge_style("bad"))

    # ========================================================================
    # STATUS HELPERS
    # ========================================================================
    @staticmethod
    def _severity_info(severity):
        return SEVERITY_CONFIG.get(severity, ("⚪", "#666666"))

    @staticmethod
    def _status_cfg(status):
        return STATUS_CONFIG.get(status, STATUS_CONFIG["NOT POWER-LAW"])

    # ========================================================================
    # DIALOG ACTIONS
    # ========================================================================
    def _show_methodology(self):
        if (
            self._methodology_dialog is None
            or not self._methodology_dialog.isVisible()
        ):
            self._methodology_dialog = MethodologyDialog(self)
            self._methodology_dialog.show()
        else:
            self._methodology_dialog.raise_()
            self._methodology_dialog.activateWindow()

    def refresh_layers(self):
        self.layer_combo.clear()
        self.layer_combo.addItem("-- Select Layer --", None)
        for layer in QgsProject.instance().mapLayers().values():
            if isinstance(layer, QgsVectorLayer):
                self.layer_combo.addItem(layer.name(), layer.id())

    def _on_layer_changed(self, idx):
        lid = self.layer_combo.currentData()
        if lid:
            self.layer = QgsProject.instance().mapLayer(lid)
            self.external_data = None
            self.file_label.setText("Using QGIS layer")
            self._update_attributes()
        else:
            self.layer = None

    def _load_file(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "Open File", "", "Data (*.csv *.xlsx)"
        )
        if path:
            try:
                if path.endswith(".csv"):
                    self.external_data = pd.read_csv(path)
                else:
                    self.external_data = pd.read_excel(path)
                self.layer = None
                self.file_label.setText(os.path.basename(path))
                self._update_attributes()
            except Exception as e:
                self._show_error(str(e))

    def _update_attributes(self):
        self.attr_combo.clear()
        self.attr_combo.addItem("-- Select Attribute --", None)

        if self.external_data is not None:
            for col in self.external_data.select_dtypes(
                include=[np.number]
            ).columns:
                self.attr_combo.addItem(col, col)
        elif self.layer:
            for field in self.layer.fields():
                if self._is_numeric_field(field):
                    self.attr_combo.addItem(field.name(), field.name())

    @staticmethod
    def _is_numeric_field(field):
        type_name = (
            field.typeName().lower() if hasattr(field, "typeName") else ""
        )
        is_numeric = (
            field.isNumeric() if hasattr(field, "isNumeric") else False
        )
        numeric_types = {
            "integer", "real", "double", "int64", "integer64",
            "float", "decimal",
        }
        return is_numeric or type_name in numeric_types

    def _on_attr_changed(self, idx):
        if self.attr_combo.currentData():
            self._update_info()
            self.run_btn.setEnabled(True)
        else:
            self.run_btn.setEnabled(False)

    def _update_info(self):
        data = self._get_data()
        attr = self.attr_combo.currentData()
        if data is not None and attr and attr in data.columns:
            vals = data[attr].dropna()
            self.info_label.setText(
                f"Samples: {len(vals):,} | "
                f"Range: [{vals.min():.4f}, {vals.max():.4f}]"
            )

    # ========================================================================
    # DATA EXTRACTION
    # ========================================================================
    def _get_data(self):
        if self.external_data is not None:
            return self.external_data
        if self.layer is not None:
            return self._extract_layer_data()
        return None

    def _extract_layer_data(self):
        fields = self.layer.fields()
        field_names = [f.name() for f in fields]
        data = {name: [] for name in field_names}

        request = QgsFeatureRequest().setFlags(QgsFeatureRequest.NoGeometry)
        for feat in self.layer.getFeatures(request):
            for name in field_names:
                try:
                    val = feat[name]
                    data[name].append(
                        float(val) if val not in [None, ""] else np.nan
                    )
                except (ValueError, TypeError):
                    data[name].append(np.nan)
        return pd.DataFrame(data)

    # ========================================================================
    # INPUT VALIDATION
    # ========================================================================
    @staticmethod
    def _validate_input(values):
        if len(values) < MIN_TOTAL_POINTS:
            raise ValueError(
                f"Need at least {MIN_TOTAL_POINTS} data points"
            )
        if np.ptp(values) < EPS:
            raise ValueError("Zero variance data — all values are identical")

    # ========================================================================
    # SCALE GENERATION
    # ========================================================================
    def _get_scale_count(self):
        mode = self.scale_mode_combo.currentText()
        if "Fast" in mode:
            return FAST_SCALES
        if "High" in mode:
            return HIGH_PRECISION_SCALES
        return STANDARD_SCALES

    @staticmethod
    def _generate_scales(data_range, num_scales):
        scales = np.logspace(-3, -0.3, num_scales) * data_range
        scales = np.unique(np.round(scales, 10))
        return sorted(scales, reverse=True)

    # ========================================================================
    # BOX COUNTING
    # ========================================================================
    @staticmethod
    def _count_occupied(values_norm, eps_norm):
        n_intervals = int(np.floor(1.0 / eps_norm))
        if n_intervals < 2:
            return 0

        # Tiny jitter prevents boundary artifacts
        jitter = np.random.uniform(-1e-12, 1e-12, len(values_norm))
        values_jittered = np.clip(values_norm + jitter, 0.0, 1.0)

        bins = np.linspace(0.0, 1.0, n_intervals + 1)
        indices = np.digitize(values_jittered, bins) - 1
        indices = np.clip(indices, 0, n_intervals - 1)
        return len(np.unique(indices))

    def _compute_box_counts(self, values_norm, scales, data_range, n_total):
        occupied = []
        valid = []
        for eps in scales:
            eps_norm = np.clip(eps / data_range, 1e-6, 0.5)
            if eps_norm <= 0 or eps_norm > 1:
                continue
            n_occ = self._count_occupied(values_norm, eps_norm)
            if 1 <= n_occ <= n_total:
                occupied.append(n_occ)
                valid.append(eps)
        return valid, occupied

    @staticmethod
    def _clean_duplicates(scales, occupied):
        clean_scales, clean_occ = [], []
        prev = None
        for eps, n_occ in zip(scales, occupied):
            if n_occ != prev or not clean_scales:
                clean_scales.append(eps)
                clean_occ.append(n_occ)
                prev = n_occ
        if len(clean_scales) < 5:
            return scales, occupied
        return clean_scales, clean_occ

    # ========================================================================
    # SCALING REGION DETECTION
    # ========================================================================
    def _find_scaling_region(self, log_eps, log_N):
        n = len(log_eps)
        if n < MIN_SCALING_POINTS:
            return list(range(n)), 0.0, 0.0, 0.0

        min_pts = max(MIN_SCALING_POINTS, n // 3)
        best = {
            "score": float("inf"),
            "start": 0,
            "end": n,
            "r2": 0.0,
            "slope": 0.0,
            "intercept": 0.0,
        }

        X, Y = np.array(log_eps), np.array(log_N)
        cx, cy = np.cumsum(X), np.cumsum(Y)
        cxy, cxx = np.cumsum(X * Y), np.cumsum(X * X)

        for start in range(n - min_pts):
            off_x = cx[start - 1] if start > 0 else 0
            off_y = cy[start - 1] if start > 0 else 0
            off_xy = cxy[start - 1] if start > 0 else 0
            off_xx = cxx[start - 1] if start > 0 else 0

            for end in range(start + min_pts, n + 1):
                k = end - start
                s_x = cx[end - 1] - off_x
                s_y = cy[end - 1] - off_y
                s_xy = cxy[end - 1] - off_xy
                s_xx = cxx[end - 1] - off_xx

                denom = k * s_xx - s_x * s_x
                if abs(denom) < EPS:
                    continue

                slope = (k * s_xy - s_x * s_y) / denom
                intercept = (s_y - slope * s_x) / k

                y_pred = slope * X[start:end] + intercept
                ss_res = np.sum((Y[start:end] - y_pred) ** 2)
                ss_tot = max(
                    EPS, np.sum((Y[start:end] - np.mean(Y[start:end])) ** 2)
                )
                r2 = 1.0 - ss_res / ss_tot if ss_tot > EPS else 0.0

                rmse = np.sqrt(ss_res / k)
                score = (
                    rmse * 0.6 + (1.0 / k) * 0.3 + (1.0 - r2) * 0.4
                )

                if score < best["score"] and r2 > R2_REGION_MIN:
                    best.update(
                        {
                            "score": score,
                            "start": start,
                            "end": end,
                            "r2": r2,
                            "slope": slope,
                            "intercept": intercept,
                        }
                    )

        if best["end"] - best["start"] < MIN_SCALING_POINTS:
            return list(range(n)), 0.0, 0.0, 0.0

        indices = list(range(best["start"], best["end"]))
        return indices, best["r2"], best["slope"], best["intercept"]

    # ========================================================================
    # REGRESSION
    # ========================================================================
    @staticmethod
    def _compute_regression(x_reg, y_reg):
        coeffs = np.polyfit(x_reg, y_reg, 1)
        slope, intercept = coeffs
        y_pred = np.polyval(coeffs, x_reg)
        ss_res = np.sum((y_reg - y_pred) ** 2)
        ss_tot = np.sum((y_reg - np.mean(y_reg)) ** 2)
        r2 = 1.0 - ss_res / ss_tot if ss_tot > EPS else 0.0
        return slope, intercept, r2, ss_res

    # ========================================================================
    # CONFIDENCE INTERVAL
    # ========================================================================
    @staticmethod
    def _compute_confidence_interval(x_reg, ss_res, D):
        n_pts = len(x_reg)
        if n_pts <= 2:
            return None

        s_xx = np.sum((x_reg - np.mean(x_reg)) ** 2)
        if s_xx < EPS:
            return None

        sigma2 = ss_res / (n_pts - 2)
        se_slope = np.sqrt(sigma2 / s_xx)
        t_val = stats.t.ppf(0.975, n_pts - 2)
        half_width = t_val * se_slope
        return (max(0.0, D - half_width), D + half_width)

    # ========================================================================
    # SURROGATE TEST
    # ========================================================================
    def _surrogate_test(self, values, D_real, n_surrogates=50):
        np.random.seed(SURROGATE_SEED)
        values = np.asarray(values, dtype=float)
        D_surrogates = []

        for _ in range(n_surrogates):
            shuffled = np.random.permutation(values)
            vmin, vmax = np.min(shuffled), np.max(shuffled)
            data_range = vmax - vmin
            if data_range < EPS:
                continue

            values_norm = (shuffled - vmin) / data_range
            scales = (
                np.logspace(-3, -0.3, SURROGATE_SCALES) * data_range
            )
            scales = sorted(scales, reverse=True)

            occupied, valid_scales = [], []
            for eps in scales:
                eps_norm = np.clip(eps / data_range, 1e-6, 0.5)
                if eps_norm <= 0 or eps_norm > 1:
                    continue
                n_occ = self._count_occupied(values_norm, eps_norm)
                if 1 <= n_occ <= len(shuffled):
                    occupied.append(n_occ)
                    valid_scales.append(eps)

            if len(occupied) >= MIN_SCALING_POINTS:
                occupied = np.maximum(occupied, 1)
                log_eps = np.log10(valid_scales)
                log_N = np.log10(occupied)
                coeffs = np.polyfit(log_eps, log_N, 1)
                D_surrogates.append(abs(coeffs[0]))

        if len(D_surrogates) < MIN_SURROGATES:
            return 1.0, np.array([])

        D_surrogates = np.array(D_surrogates)
        p_value = np.mean(np.abs(D_surrogates) >= np.abs(D_real))
        p_value = max(1.0 / len(D_surrogates), min(1.0, p_value))
        return p_value, D_surrogates

    # ========================================================================
    # DIAGNOSTICS
    # ========================================================================
    def _compute_diagnostics(
        self,
        log_eps,
        log_N,
        r2,
        n_total,
        data_std,
        d_ci=None,
        surrogate_p=None,
        n_total_scales=None,
        n_region=None,
    ):
        warnings = []
        diag = {"n_points": len(log_eps), "r2": r2}

        diag["data_quality_ok"] = (
            n_total >= MIN_SAMPLE_SIZE
            and data_std > 0
            and len(log_eps) >= MIN_SCALING_POINTS
        )

        if not diag["data_quality_ok"] and n_total < MIN_SAMPLE_SIZE:
            warnings.append(
                (f"Small sample size ({n_total} points)", SEVERITY_WARNING)
            )

        diag["scale_span"] = log_eps.max() - log_eps.min()

        scaling_fraction = (
            n_region / n_total_scales
            if (n_total_scales and n_region)
            else 1.0
        )
        diag["scaling_fraction"] = scaling_fraction

        self._check_scale_span(diag["scale_span"], r2, warnings)
        self._check_curvature(log_eps, log_N, diag, warnings)
        self._check_slope_stability(log_eps, log_N, diag, warnings)
        self._check_ci(d_ci, diag, warnings)
        self._check_surrogate(surrogate_p, diag, warnings)
        self._check_r2(r2, warnings)
        self._check_scaling_fraction(scaling_fraction, warnings)

        diag["confidence"] = self._compute_confidence(
            r2, diag.get("slope_std", 1.0), diag.get("curvature", 1.0)
        )

        diag["marginal_power_law"] = (
            CONFIDENCE_MARGINAL_LOW
            <= diag["confidence"]
            < CONFIDENCE_MARGINAL_HIGH
        ) or (
            diag["scale_span"] > SCALE_SPAN_MIN
            and diag.get("slope_std", 1.0) > SLOPE_UNSTABLE
            and diag["confidence"] >= 0.55
        )

        diag["is_valid_power_law"] = (
            r2 > R2_GOOD
            and diag["confidence"] > CONFIDENCE_STRONG
            and scaling_fraction > 0.4
            and diag["scale_span"] > SCALE_SPAN_MIN
            and diag["data_quality_ok"]
            and (surrogate_p is None or surrogate_p < SURROGATE_SIGNIFICANCE)
        )

        diag["status"] = self._determine_status(diag)
        return diag["status"], diag["confidence"], warnings, diag

    def _check_scale_span(self, scale_span, r2, warnings):
        if scale_span < SCALE_SPAN_MIN:
            warnings.append(
                (
                    f"Narrow scaling range ({scale_span:.2f} decades)",
                    SEVERITY_CRITICAL,
                )
            )
        elif scale_span < SCALE_SPAN_WEAK:
            warnings.append(
                (
                    f"Weak scaling range (<{SCALE_SPAN_WEAK} decades)",
                    SEVERITY_WARNING,
                )
            )

        if scale_span < SCALE_SPAN_MIN and r2 > R2_EXCELLENT:
            warnings.append(
                (
                    "High R² on narrow range → possible pseudo power-law",
                    SEVERITY_CRITICAL,
                )
            )

    def _check_curvature(self, log_eps, log_N, diag, warnings):
        n = len(log_eps)
        if n >= MIN_SCALING_POINTS:
            x, y = np.array(log_eps), np.array(log_N)
            dx = np.diff(x)
            d2y = np.diff(np.diff(y))
            curvature = np.nanmean(np.abs(d2y) / (dx[:-1] + EPS))
            diag["curvature"] = curvature
            if curvature > CURVATURE_HIGH:
                warnings.append(
                    (f"High curvature ({curvature:.3f})", SEVERITY_CRITICAL)
                )
            elif curvature > CURVATURE_MODERATE:
                warnings.append(
                    (
                        f"Moderate curvature ({curvature:.3f})",
                        SEVERITY_WARNING,
                    )
                )
        else:
            diag["curvature"] = 1.0

    def _check_slope_stability(self, log_eps, log_N, diag, warnings):
        n = len(log_eps)
        if n >= 3:
            local_slopes = []
            for i in range(n - 2):
                xw = np.array(log_eps[i : i + 3])
                yw = np.array(log_N[i : i + 3])
                if np.std(xw) > EPS:
                    try:
                        local_slopes.append(np.polyfit(xw, yw, 1)[0])
                    except (ValueError, np.linalg.LinAlgError):
                        pass
            slope_std = (
                np.std(local_slopes) if len(local_slopes) > 1 else 1.0
            )
            diag["slope_std"] = slope_std
            if slope_std > SLOPE_UNSTABLE:
                warnings.append(
                    (
                        f"Unstable slope (σ={slope_std:.3f})",
                        SEVERITY_CRITICAL,
                    )
                )
            elif slope_std > SLOPE_WARNING:
                warnings.append(
                    (
                        f"Slightly unstable slope (σ={slope_std:.3f})",
                        SEVERITY_WARNING,
                    )
                )
        else:
            diag["slope_std"] = 1.0

    @staticmethod
    def _check_ci(d_ci, diag, warnings):
        if d_ci is not None:
            diag["D_CI_lower"], diag["D_CI_upper"] = d_ci
            ci_width = d_ci[1] - d_ci[0]
            if ci_width > CI_WIDE:
                warnings.append(
                    (
                        f"Wide confidence interval (width={ci_width:.3f})",
                        SEVERITY_WARNING,
                    )
                )

    @staticmethod
    def _check_surrogate(surrogate_p, diag, warnings):
        if surrogate_p is not None:
            diag["surrogate_p_value"] = surrogate_p
            if surrogate_p > SURROGATE_SIGNIFICANCE:
                warnings.append(
                    (
                        f"Surrogate test NOT significant (p={surrogate_p:.3f})",
                        SEVERITY_CRITICAL,
                    )
                )
            elif surrogate_p > SURROGATE_MARGINAL:
                warnings.append(
                    (
                        f"Surrogate test marginally significant (p={surrogate_p:.3f})",
                        SEVERITY_WARNING,
                    )
                )
            else:
                warnings.append(
                    (
                        f"Surrogate test significant (p={surrogate_p:.3f}) ✓",
                        SEVERITY_OK,
                    )
                )

    @staticmethod
    def _check_r2(r2, warnings):
        if r2 < R2_ACCEPTABLE:
            warnings.append((f"Low R² ({r2:.3f})", SEVERITY_CRITICAL))
        elif r2 < R2_GOOD:
            warnings.append((f"Moderate R² ({r2:.3f})", SEVERITY_WARNING))
        elif r2 < R2_EXCELLENT:
            warnings.append((f"Acceptable R² ({r2:.3f})", SEVERITY_INFO))

    @staticmethod
    def _check_scaling_fraction(fraction, warnings):
        if fraction < SCALING_FRACTION_CRITICAL:
            warnings.append(
                (
                    f"Very small scaling region ({fraction:.1%})",
                    SEVERITY_CRITICAL,
                )
            )
        elif fraction < SCALING_FRACTION_MIN:
            warnings.append(
                (
                    f"Small scaling region ({fraction:.1%})",
                    SEVERITY_WARNING,
                )
            )

    @staticmethod
    def _compute_confidence(r2, slope_std, curvature):
        r2_score = max(0.0, min(1.0, r2))
        return (
            0.5 * r2_score
            + 0.3 * np.exp(-slope_std)
            + 0.2 * np.exp(-curvature)
        )

    def _determine_status(self, diag):
        if diag.get("is_valid_power_law"):
            return "STRONG POWER-LAW"
        if diag["confidence"] >= CONFIDENCE_WEAK and diag["data_quality_ok"]:
            return "WEAK POWER-LAW"
        if diag.get("marginal_power_law") and diag["data_quality_ok"]:
            return "MARGINAL POWER-LAW"
        return "NOT POWER-LAW"

    # ========================================================================
    # MAIN ANALYSIS PIPELINE
    # ========================================================================
    def run_analysis(self):
        """Execute the complete fractal dimension analysis workflow."""
        data = self._get_data()
        attr = self.attr_combo.currentData()

        if data is None or not attr:
            return

        try:
            self.timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            self._set_analysis_state(running=True)

            values = data[attr].dropna().values
            self._validate_input(values)

            vmin, vmax = np.min(values), np.max(values)
            data_range = vmax - vmin
            values_norm = (values - vmin) / data_range

            self._update_progress(15)

            num_scales = self._get_scale_count()
            scales = self._generate_scales(data_range, num_scales)

            valid_scales, occupied = self._compute_box_counts(
                values_norm, scales, data_range, len(values)
            )

            if len(valid_scales) < MIN_SCALING_POINTS:
                raise ValueError(
                    f"Only {len(valid_scales)} valid scales found."
                )

            valid_scales, occupied = self._clean_duplicates(
                valid_scales, occupied
            )
            occupied = np.maximum(occupied, 1)

            if len(valid_scales) < MIN_VALID_SCALES:
                raise ValueError(
                    f"Insufficient scale resolution "
                    f"({len(valid_scales)} scales). Need ≥ {MIN_VALID_SCALES}."
                )

            log_eps = np.log10(valid_scales)
            log_N = np.log10(occupied)

            self._update_progress(30)

            scaling_idx, _, _, _ = self._find_scaling_region(log_eps, log_N)

            self._update_progress(50)

            x_reg, y_reg = log_eps[scaling_idx], log_N[scaling_idx]
            slope, intercept, r2, ss_res = self._compute_regression(
                x_reg, y_reg
            )
            D = abs(slope)

            d_ci = self._compute_confidence_interval(x_reg, ss_res, D)

            self._update_progress(70)

            self._status_label.setText("Running surrogate test...")
            QApplication.processEvents()
            surrogate_p, surrogate_D = self._surrogate_test(values, D)

            self._update_progress(90)

            status, confidence, warnings, diagnostics = (
                self._compute_diagnostics(
                    log_eps[scaling_idx],
                    log_N[scaling_idx],
                    r2,
                    len(values),
                    np.std(values),
                    d_ci=d_ci,
                    surrogate_p=surrogate_p,
                    n_total_scales=len(valid_scales),
                    n_region=len(scaling_idx),
                )
            )

            self.results = self._build_results_dict(
                attr,
                values,
                valid_scales,
                occupied.tolist(),
                scaling_idx,
                slope,
                intercept,
                r2,
                D,
                d_ci,
                status,
                confidence,
                warnings,
                diagnostics,
                surrogate_p,
                num_scales,
            )

            self._update_progress(100)

            self._update_badges(
                len(values), diagnostics.get("scale_span", 0), len(valid_scales)
            )
            self.display_results()
            self.update_plot()

            self._set_export_buttons(True)
            self._show_alerts(diagnostics, status, confidence, surrogate_p)

        except Exception as e:
            self._set_analysis_state(running=False)
            self._status_label.setText("Error")
            self._status_label.setStyleSheet("color: red;")
            self._show_error(str(e))

    def _set_analysis_state(self, running):
        self.run_btn.setEnabled(not running)
        self._progress_bar.setVisible(running)
        if running:
            self._progress_bar.setValue(0)

    def _update_progress(self, value):
        self._progress_bar.setValue(value)
        QApplication.processEvents()

    def _set_export_buttons(self, enabled):
        buttons = [
            self.export_btn,
            self.export_png_btn,
            self.export_jpg_btn,
            self.export_pdf_btn,
            self.export_clipboard_btn,
            self.update_plot_btn,
        ]
        for btn in buttons:
            btn.setEnabled(enabled)

    def _show_alerts(self, diagnostics, status, confidence, surrogate_p):
        if not diagnostics.get("data_quality_ok", False):
            self._show_warning(
                "Data Quality Alert",
                f"Insufficient data quality — "
                f"sample size: {self.results['samples']} "
                f"(recommended ≥ {MIN_SAMPLE_SIZE})",
            )
        elif status != "STRONG POWER-LAW":
            self._show_warning(
                "Fractal Analysis Caution",
                f"Status: {status}\n"
                f"Confidence: {confidence:.2f}\n"
                f"Surrogate p: {surrogate_p:.3f}",
            )

    # ========================================================================
    # RESULTS BUILDER
    # ========================================================================
    def _build_results_dict(
        self,
        attr,
        values,
        valid_scales,
        occupied,
        scaling_idx,
        slope,
        intercept,
        r2,
        D,
        d_ci,
        status,
        confidence,
        warnings,
        diagnostics,
        surrogate_p,
        num_scales,
    ):
        eps_dense = None
        N_regression = None
        if len(scaling_idx) > 1:
            eps_min = valid_scales[scaling_idx[-1]]
            eps_max = valid_scales[scaling_idx[0]]
            eps_dense = np.logspace(
                np.log10(eps_min), np.log10(eps_max), 100
            )
            N_regression = 10 ** (
                slope * np.log10(eps_dense) + intercept
            )

        return {
            "attribute": attr,
            "samples": len(values),
            "range": (np.min(values), np.max(values)),
            "D": D,
            "D_CI": d_ci,
            "R2": r2,
            "slope": slope,
            "intercept": intercept,
            "scales": valid_scales,
            "occupied": occupied,
            "log_eps": np.log10(valid_scales).tolist(),
            "log_N": np.log10(np.maximum(occupied, 1)).tolist(),
            "scaling_idx": scaling_idx,
            "region_points": len(scaling_idx),
            "total_scales": len(valid_scales),
            "eps_regression": eps_dense,
            "N_regression": N_regression,
            "power_law_status": status,
            "power_law_confidence": confidence,
            "power_law_warnings": warnings,
            "diagnostics": diagnostics,
            "surrogate_p": surrogate_p,
            "timestamp": self.timestamp,
            "method": PLUGIN_METHOD,
            "version": PLUGIN_VERSION,
            "definition": (
                "box-counting empirical scaling exponent over "
                "selected region"
            ),
            "interpretation_note": (
                "D is empirical estimator, not invariant fractal constant"
            ),
            "num_scales": num_scales,
        }

    # ========================================================================
    # DISPLAY RESULTS
    # ========================================================================
    def display_results(self):
        if not self.results:
            return

        r = self.results
        self.results_table.setRowCount(0)
        diag = r.get("diagnostics", {})
        data_ok = diag.get("data_quality_ok", False)

        self._add_quality_warning(r, data_ok)
        self._add_powerlaw_warnings(r)
        self._add_status_row(r)
        self._add_result_rows(r, diag)
        self._update_status_label(r, diag, data_ok)

        self.results_table.resizeColumnsToContents()
        self._progress_bar.setVisible(False)
        self.run_btn.setEnabled(True)

    def _add_quality_warning(self, r, data_ok):
        if not data_ok:
            row = self.results_table.rowCount()
            self.results_table.insertRow(row)
            item = QTableWidgetItem("⚠ DATA QUALITY")
            item.setBackground(Qt.red)
            item.setForeground(Qt.white)
            self.results_table.setItem(row, 0, item)
            self.results_table.setItem(
                row, 1, QTableWidgetItem(f"Small sample (N={r['samples']})")
            )

    def _add_powerlaw_warnings(self, r):
        for msg, severity in r.get("power_law_warnings", []):
            row = self.results_table.rowCount()
            self.results_table.insertRow(row)
            icon, _ = self._severity_info(severity)
            item = QTableWidgetItem(f"{icon} {msg}")
            if severity >= SEVERITY_CRITICAL:
                item.setBackground(QColor("#ffe6e6"))
            elif severity >= SEVERITY_WARNING:
                item.setBackground(QColor("#fff8e6"))
            self.results_table.setItem(row, 0, item)
            self.results_table.setItem(row, 1, QTableWidgetItem(""))

    def _add_status_row(self, r):
        row = self.results_table.rowCount()
        self.results_table.insertRow(row)
        item = QTableWidgetItem("Status")
        cfg = self._status_cfg(r["power_law_status"])
        item.setBackground(cfg["bg"])
        if r["power_law_status"] == "NOT POWER-LAW":
            item.setForeground(Qt.white)
        self.results_table.setItem(row, 0, item)
        self.results_table.setItem(
            row,
            1,
            QTableWidgetItem(
                f"{r['power_law_status']} "
                f"(conf: {r['power_law_confidence']:.2f})"
            ),
        )

    def _add_result_rows(self, r, diag):
        d_text = f"{r['D']:.4f}"
        if r.get("D_CI"):
            d_text += (
                f"  [95% CI: {r['D_CI'][0]:.4f}, {r['D_CI'][1]:.4f}]"
            )

        items = [
            ("Definition", r.get("definition", "")),
            ("Fractal Dimension D", d_text),
            ("R²", f"{r['R2']:.4f}"),
            ("Surrogate p-value", f"{r.get('surrogate_p', 'N/A'):.4f}"),
            (
                "Scale span",
                f"{diag.get('scale_span', 0):.2f} decades",
            ),
            (
                "Scaling fraction",
                f"{diag.get('scaling_fraction', 0):.1%}",
            ),
            ("Curvature", f"{diag.get('curvature', 0):.4f}"),
            ("Slope stability σ", f"{diag.get('slope_std', 0):.4f}"),
            (
                "Region points / Total",
                f"{r['region_points']} / {r['total_scales']}",
            ),
            ("Analysis scales", str(r.get("num_scales", STANDARD_SCALES))),
            ("Method", PLUGIN_METHOD),
            ("Version", PLUGIN_VERSION),
        ]

        for param, val in items:
            row = self.results_table.rowCount()
            self.results_table.insertRow(row)
            item = QTableWidgetItem(param)
            if param == "Fractal Dimension D":
                item.setToolTip(
                    "Box-counting empirical scaling exponent. "
                    "Not an invariant fractal constant."
                )
            self.results_table.setItem(row, 0, item)
            self.results_table.setItem(row, 1, QTableWidgetItem(val))

    def _update_status_label(self, r, diag, data_ok):
        cfg = self._status_cfg(r["power_law_status"])

        if not data_ok:
            self._status_label.setStyleSheet(
                "font-weight: bold; color: white; "
                "background-color: #cc0000; padding: 4px; font-size: 11px;"
            )
        else:
            self._status_label.setStyleSheet(cfg["style"])

        d_ci_str = ""
        if r.get("D_CI"):
            d_ci_str = f" [{r['D_CI'][0]:.3f}, {r['D_CI'][1]:.3f}]"

        n_crit = sum(
            1
            for _, s in r.get("power_law_warnings", [])
            if s >= SEVERITY_CRITICAL
        )
        crit_str = f" | ⚠{n_crit}" if n_crit > 0 else ""

        self._status_label.setText(
            f"{cfg['icon']} D={r['D']:.3f}{d_ci_str} | "
            f"R²={r['R2']:.3f} | {r['power_law_status']}{crit_str}"
        )

    # ========================================================================
    # PLOT
    # ========================================================================
    def update_plot(self):
        if not self.results or not self.results.get("scales"):
            return

        self.figure.clear()
        r = self.results
        eps = np.array(r["scales"])
        N_occ = np.array(r["occupied"])
        idx = r["scaling_idx"]
        diag = r.get("diagnostics", {})

        ax = self.figure.add_subplot(111)
        self.figure.subplots_adjust(
            left=0.13, right=0.94, top=0.92, bottom=0.13
        )

        self._draw_scaling_window(ax, eps, idx, N_occ)
        self._draw_points(ax, eps, N_occ, idx)
        self._draw_regression(ax, r)
        self._draw_confidence_band(ax, r)
        self._draw_labels(ax, eps, N_occ, idx)
        self._draw_info_box(ax, r, diag)
        self._format_axes(ax, r, eps, N_occ)

        self.canvas.draw()

    @staticmethod
    def _draw_scaling_window(ax, eps, idx, N_occ):
        if len(idx) > 1:
            rect = Rectangle(
                (eps[idx[-1]], 0),
                eps[idx[0]] - eps[idx[-1]],
                max(N_occ) * 1.5,
                facecolor="red",
                alpha=0.06,
                zorder=0,
            )
            ax.add_patch(rect)

    def _draw_points(self, ax, eps, N_occ, idx):
        mask = np.ones(len(eps), dtype=bool)
        mask[idx] = False

        if self.show_all_points_check.isChecked() and np.any(mask):
            ax.semilogx(
                eps[mask],
                N_occ[mask],
                "bo",
                markersize=10,
                markerfacecolor="blue",
                markeredgecolor="black",
                markeredgewidth=1.5,
                alpha=0.7,
                label="Other scales",
                zorder=2,
            )

        if len(idx) > 0:
            ax.semilogx(
                eps[idx],
                N_occ[idx],
                "rs",
                markersize=12,
                markerfacecolor="red",
                markeredgecolor="darkred",
                markeredgewidth=2,
                label=f"Scaling region (N={len(idx)})",
                zorder=3,
            )
            if len(idx) > 1:
                ax.semilogx(
                    eps[idx],
                    N_occ[idx],
                    "r-",
                    linewidth=2.5,
                    alpha=0.8,
                    zorder=2.5,
                )

    def _draw_regression(self, ax, r):
        if (
            self.show_regression_check.isChecked()
            and r["eps_regression"] is not None
        ):
            ax.semilogx(
                r["eps_regression"],
                r["N_regression"],
                "k--",
                linewidth=3,
                alpha=0.9,
                label=f'D={r["D"]:.3f}, R²={r["R2"]:.3f}',
                zorder=4,
            )

    def _draw_confidence_band(self, ax, r):
        if not (
            self.show_regression_check.isChecked()
            and r.get("D_CI")
            and r["D_CI"][1] > r["D_CI"][0]
        ):
            return

        se = (r["D_CI"][1] - r["D_CI"][0]) / 3.92
        slope_upper = r["slope"] + 1.96 * se
        slope_lower = r["slope"] - 1.96 * se

        N_upper = 10 ** (
            slope_upper * np.log10(r["eps_regression"]) + r["intercept"]
        )
        N_lower = 10 ** (
            slope_lower * np.log10(r["eps_regression"]) + r["intercept"]
        )

        ax.fill_between(
            r["eps_regression"],
            N_lower,
            N_upper,
            alpha=0.12,
            color="gray",
            label="95% CI",
        )

    def _draw_labels(self, ax, eps, N_occ, idx):
        if not self.show_labels_check.isChecked():
            return

        offset = max(N_occ) * 0.03
        for i in range(len(eps)):
            if (
                len(eps) > MAX_LABELS_BEFORE_SKIP
                and i % 2 != 0
                and i not in idx
            ):
                continue
            color = "#8b0000" if i in idx else "#2a2a8b"
            fw = "bold" if i in idx else "normal"
            fs = LABEL_FONT_REGION if i in idx else LABEL_FONT_OTHER
            ax.annotate(
                f"{N_occ[i]}",
                xy=(eps[i], N_occ[i]),
                xytext=(eps[i], N_occ[i] + offset),
                fontsize=fs,
                fontweight=fw,
                color=color,
                ha="center",
                va="bottom",
                bbox=dict(
                    boxstyle="round,pad=0.15",
                    facecolor="white",
                    edgecolor="none",
                    alpha=0.7,
                ),
            )

    def _draw_info_box(self, ax, r, diag):
        if len(r["scaling_idx"]) <= 1:
            return

        cfg = self._status_cfg(
            r.get("power_law_status", "NOT POWER-LAW")
        )

        d_str = f'D = {r["D"]:.3f}'
        if r.get("D_CI"):
            d_str += (
                f' [{r["D_CI"][0]:.3f}, {r["D_CI"][1]:.3f}]'
            )

        n_crit = sum(
            1
            for _, s in r.get("power_law_warnings", [])
            if s >= SEVERITY_CRITICAL
        )
        crit_str = f"\n⚠{n_crit} critical" if n_crit > 0 else ""

        textstr = (
            f"{d_str}\n"
            f'R² = {r["R2"]:.3f}\n'
            f'p = {r.get("surrogate_p", 0):.3f}\n'
            f'{cfg["icon"]} {r.get("power_law_status", "Unknown")}'
            f"{crit_str}"
        )

        box_color = (
            "lightcoral"
            if not diag.get("data_quality_ok", False)
            else cfg["box_color"]
        )

        ax.text(
            0.98,
            0.97,
            textstr,
            transform=ax.transAxes,
            fontsize=9,
            verticalalignment="top",
            horizontalalignment="right",
            bbox=dict(
                boxstyle="round,pad=0.4",
                facecolor=box_color,
                edgecolor="black",
                alpha=0.9,
            ),
            fontfamily="monospace",
        )

    def _format_axes(self, ax, r, eps, N_occ):
        ax.set_xlabel("ε (scale)", fontsize=12, fontweight="bold", labelpad=10)
        ax.set_ylabel(
            "N (occupied intervals)",
            fontsize=12,
            fontweight="bold",
            labelpad=10,
        )

        title = f'GeoFracBox1D - {r["attribute"]}'
        if r.get("power_law_status"):
            title += f' | {r["power_law_status"]}'
        ax.set_title(title, fontsize=12, fontweight="bold", pad=15)

        if self.show_grid_check.isChecked():
            ax.grid(
                True,
                which="major",
                linestyle="-",
                alpha=0.5,
                linewidth=0.7,
            )
            ax.grid(
                True,
                which="minor",
                linestyle="--",
                alpha=0.25,
                linewidth=0.4,
            )

        ax.legend(
            loc="upper left",
            fontsize=9,
            framealpha=0.85,
            edgecolor="black",
            fancybox=True,
        )

        if len(eps) > 0:
            ax.set_xlim(eps[-1] * 0.6, eps[0] * 1.8)
        if len(N_occ) > 0:
            ax.set_ylim(
                max(0, min(N_occ) - 0.5), max(N_occ) * 1.15
            )

        ax.tick_params(axis="both", which="major", labelsize=10)

    # ========================================================================
    # EXPORT HELPERS
    # ========================================================================
    def _get_export_meta(self, r, diag):
        """Build metadata dictionary for export."""
        meta = {
            "method": PLUGIN_METHOD,
            "version": PLUGIN_VERSION,
            "timestamp": str(r.get("timestamp", "")),
            "attribute": str(r["attribute"]),
            "definition": str(r.get("definition", "")),
            "D": round(float(r["D"]), 6),
            "R2": round(float(r["R2"]), 6),
            "status": str(r["power_law_status"]),
            "confidence": round(float(r["power_law_confidence"]), 6),
            "scale_span_decades": round(
                float(diag.get("scale_span", 0)), 6
            ),
            "scaling_fraction": round(
                float(diag.get("scaling_fraction", 0)), 6
            ),
            "curvature": round(float(diag.get("curvature", 0)), 6),
            "slope_stability_sigma": round(
                float(diag.get("slope_std", 0)), 6
            ),
            "surrogate_p_value": (
                round(float(r.get("surrogate_p", 1.0)), 6)
                if r.get("surrogate_p") is not None
                else None
            ),
            "n_samples": int(r["samples"]),
            "n_scales": int(r["total_scales"]),
            "n_region_points": int(r["region_points"]),
            "analysis_scales": int(r.get("num_scales", STANDARD_SCALES)),
            "interpretation_note": str(r.get("interpretation_note", "")),
        }

        if r.get("D_CI"):
            meta["D_CI_lower"] = round(float(r["D_CI"][0]), 6)
            meta["D_CI_upper"] = round(float(r["D_CI"][1]), 6)

        for i, (msg, severity) in enumerate(
            r.get("power_law_warnings", [])
        ):
            meta[f"warning_{i + 1}"] = f"[severity_{severity}] {msg}"

        return meta

    def _get_base_dataframe(self, r):
        """Build base DataFrame with scaling data."""
        return pd.DataFrame({
            "epsilon": r["scales"],
            "N_occupied": r["occupied"],
            "log_eps": r["log_eps"],
            "log_N": r["log_N"],
            "in_region": [
                i in r["scaling_idx"]
                for i in range(len(r["scales"]))
            ],
        })

    def _export_excel(self, path, base_df, meta):
        """Export to Excel with metadata in separate sheet."""
        base_df.to_excel(path, sheet_name="scaling_data", index=False)
        try:
            from openpyxl import load_workbook
            wb = load_workbook(path)
            ws = wb.create_sheet("metadata")
            for i, (key, val) in enumerate(meta.items(), 1):
                ws.cell(row=i, column=1, value=key)
                ws.cell(row=i, column=2, value=val)
            wb.save(path)
        except ImportError:
            meta_path = path.replace(".xlsx", "_metadata.csv")
            pd.DataFrame([meta]).to_csv(meta_path, index=False)

    def _export_csv(self, path, base_df, meta):
        """Export to CSV with metadata in a separate file."""
        # Save scaling data
        base_df.to_csv(path, index=False, float_format="%.6f")
        
        # Save metadata in a separate CSV file
        meta_path = path.replace(".csv", "_metadata.csv")
        pd.DataFrame([meta]).to_csv(meta_path, index=False)

    @staticmethod
    def _export_json(path, base_df, meta):
        """Export to JSON with metadata and scaling data."""
        out = {
            "metadata": meta,
            "scaling_data": base_df.to_dict(orient="records"),
        }
        with open(path, "w", encoding="utf-8") as f:
            json.dump(out, f, indent=2, ensure_ascii=False, default=str)

    def _save_png(self, path):
        """Save plot as PNG."""
        self.figure.savefig(
            path,
            dpi=PLOT_DPI,
            bbox_inches="tight",
            facecolor="white",
            edgecolor="none",
        )

    def _save_jpg(self, path):
        """Save plot as JPEG."""
        self.figure.savefig(
            path,
            dpi=PLOT_DPI,
            bbox_inches="tight",
            facecolor="white",
            edgecolor="none",
            format="jpg",
        )

    def _save_pdf(self, path):
        """Save comprehensive PDF report with plot."""
        from matplotlib.backends.backend_pdf import PdfPages

        with PdfPages(path) as pdf:
            # Cover page with diagnostics
            fig = plt.figure(figsize=(8.5, 11))
            r = self.results
            diag = r.get("diagnostics", {})

            fig.text(
                0.5, 0.97,
                "GeoFracBox1D Fractal Analysis Report",
                ha="center", fontsize=16, fontweight="bold",
            )
            fig.text(
                0.5, 0.94,
                f"Method: {PLUGIN_METHOD} | v{PLUGIN_VERSION} | "
                f"Timestamp: {r.get('timestamp', 'N/A')}",
                ha="center", fontsize=9, color="gray",
            )

            d_str = f"D = {r['D']:.4f}"
            if r.get("D_CI"):
                d_str += f" (95% CI: {r['D_CI'][0]:.4f}–{r['D_CI'][1]:.4f})"

            info = (
                f"Attribute: {r['attribute']}\n"
                f"Definition: {r.get('definition', '')}\n"
                f"{d_str}\n"
                f"R² = {r['R2']:.4f}\n"
                f"Status: {r.get('power_law_status', 'Unknown')}\n"
                f"Confidence: {r.get('power_law_confidence', 0):.2f}\n"
                f"Surrogate p: {r.get('surrogate_p', 'N/A'):.4f}\n\n"
                f"Scale span: {diag.get('scale_span', 0):.2f} decades\n"
                f"Scaling fraction: {diag.get('scaling_fraction', 0):.1%}\n"
                f"Curvature: {diag.get('curvature', 0):.4f}\n"
                f"Slope stability σ: {diag.get('slope_std', 0):.4f}\n"
                f"Analysis scales: {r.get('num_scales', STANDARD_SCALES)}\n\n"
                f"Warnings ({len(r.get('power_law_warnings', []))}):\n"
            )

            for msg, sev in r.get("power_law_warnings", []):
                icon, _ = self._severity_info(sev)
                info += f"  [{icon}] {msg}\n"

            info += f"\nNote: {r.get('interpretation_note', '')}"

            fig.text(
                0.5, 0.40, info,
                ha="center", fontsize=10, fontfamily="monospace",
                verticalalignment="top",
            )

            pdf.savefig(fig)
            plt.close(fig)

            # Plot page
            pdf.savefig(self.figure)

    # ========================================================================
    # PUBLIC EXPORT METHODS
    # ========================================================================
    def copy_to_clipboard(self):
        """Copy analysis results summary to clipboard."""
        if not self.results:
            self._show_warning("No Data", "Run analysis first before copying.")
            return

        r = self.results
        diag = r.get("diagnostics", {})

        d_str = f"D = {r['D']:.4f}"
        if r.get("D_CI"):
            d_str += f" (95% CI: {r['D_CI'][0]:.4f}–{r['D_CI'][1]:.4f})"

        warnings_str = "\n".join(
            f"  [{self._severity_info(s)[0]}] {m}"
            for m, s in r.get("power_law_warnings", [])
        )

        summary = f"""GeoFracBox1D v{PLUGIN_VERSION} - Fractal Analysis Results
{'=' * 60}
Attribute: {r['attribute']}
Timestamp: {r.get('timestamp', 'N/A')}
Analysis scales: {r.get('num_scales', STANDARD_SCALES)}

CORE RESULTS:
  Definition: {r.get('definition', '')}
  {d_str}
  R² = {r['R2']:.4f}
  Status: {r['power_law_status']}
  Confidence: {r['power_law_confidence']:.2f}

DIAGNOSTICS:
  Scale span: {diag.get('scale_span', 0):.2f} decades
  Scaling fraction: {diag.get('scaling_fraction', 0):.1%}
  Curvature: {diag.get('curvature', 0):.4f}
  Slope stability σ: {diag.get('slope_std', 0):.4f}
  Surrogate p-value: {r.get('surrogate_p', 'N/A'):.4f}
  Data quality OK: {diag.get('data_quality_ok', False)}

WARNINGS:
{warnings_str if warnings_str else '  None'}

Note: {r.get('interpretation_note', '')}"""

        QApplication.clipboard().setText(summary.strip())
        self._show_info("Copied", "Results summary copied to clipboard!")

    def export_result(self):
        """Export results in selected format (Excel/CSV/JSON)."""
        if not self.results:
            self._show_warning("No Data", "Run analysis first before exporting.")
            return

        r = self.results
        diag = r.get("diagnostics", {})
        fmt = self.export_format_combo.currentText()

        format_map = {
            "Excel (.xlsx)": ("geofracbox1d_results.xlsx", "Excel Files (*.xlsx)"),
            "CSV (.csv)": ("geofracbox1d_results.csv", "CSV Files (*.csv)"),
            "JSON (.json)": ("geofracbox1d_results.json", "JSON Files (*.json)"),
        }

        default_name, file_filter = format_map.get(
            fmt, ("geofracbox1d_export", "All Files (*.*)")
        )
        path, _ = QFileDialog.getSaveFileName(
            self, "Export Results", default_name, file_filter
        )

        if not path:
            return

        try:
            base_df = self._get_base_dataframe(r)
            meta = self._get_export_meta(r, diag)

            export_map = {
                "Excel (.xlsx)": lambda: self._export_excel(path, base_df, meta),
                "CSV (.csv)": lambda: self._export_csv(path, base_df, meta),
                "JSON (.json)": lambda: self._export_json(path, base_df, meta),
            }

            export_fn = export_map.get(fmt)
            if export_fn:
                export_fn()

            self._show_info("Export Successful", f"Results exported to:\n{path}")
        except Exception as e:
            self._show_error(f"Failed to export results:\n{str(e)}")

    def _export_plot(self, fmt):
        """Export plot in specified format (PNG/JPG/PDF)."""
        if not self.results:
            self._show_warning("No Data", "Run analysis first before exporting plot.")
            return

        plot_map = {
            "png": ("Save PNG", "geofracbox1d_plot.png", "PNG (*.png)", self._save_png),
            "jpg": ("Save JPG", "geofracbox1d_plot.jpg", "JPEG (*.jpg *.jpeg)", self._save_jpg),
            "pdf": ("Save PDF", "geofracbox1d_report.pdf", "PDF (*.pdf)", self._save_pdf),
        }

        info = plot_map.get(fmt)
        if not info:
            return

        title, default_name, file_filter, save_fn = info
        path, _ = QFileDialog.getSaveFileName(self, title, default_name, file_filter)

        if path:
            try:
                save_fn(path)
                self._show_info("Done", f"Plot saved to:\n{path}")
            except Exception as e:
                self._show_error(f"Failed to export plot:\n{str(e)}")


# Run the dialog
dialog = FractalAnalysisDialog()